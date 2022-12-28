import asyncio
import datetime
from aiogram import Bot, executor, types, Dispatcher
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from db import Database
from openpyxl import load_workbook


bot = Bot(token="5394783980:AAFmlnp_lWi3wzkFTNUhp1fKbv9btLB4-Bk")
storage = MemoryStorage()
dp = Dispatcher(bot, storage=storage)
db = Database('db.db')


async def main1():
    while True:
        timing = datetime.datetime.now()

        if timing.strftime("%H:%M") == "9:15":

            wb = load_workbook(filename=r'C:\Users\User\Desktop\НИР\moygrafik.xlsx')
            sheet_ranges = wb['TDSheet']
            table_numbers = [str(table_number.value) for table_number in sheet_ranges["B"]][1:]

            for table_number in table_numbers:
                try:
                    tg_id = "".join(db.tgIdSelector(table_number)[0])
                    await bot.send_message(tg_id, "Вы не отметили приход в системе MoyGrafik")
                except:
                    pass
            break



async def main2():
    while True:
        timing = datetime.datetime.now()

        if timing.strftime("%H:%M") == "18:00":

            wb = load_workbook(filename=r'C:\Users\User\Desktop\НИР\moygrafik.xlsx')  # сюда указывается имя файла
            sheet_ranges = wb['TDSheet']  # сюда указывается имя листа из экселя
            table_numbers = [str(table_number.value) for table_number in sheet_ranges["B"]][1:]

            for table_number in table_numbers:
                try:
                    tg_id = "".join(db.tgIdSelector(table_number)[0])

                    await bot.send_message(tg_id, "Вы не отметили уход в системе MoyGrafik")
                except:
                    pass
            break


if __name__ == '__main__':
    loop = asyncio.get_event_loop()

    try:
        loop.run_until_complete(main1())
    except Exception:
        pass
    try:
        loop.run_until_complete(main2())
    except Exception:
        pass
